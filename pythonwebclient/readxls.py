import openpyxl
path = "C:/Users/Administrator/PycharmProjects/pythonwebclient/data.xlsx";
# create the object of the work book
wb_obj = openpyxl.load_workbook(path)
sheet_obj = wb_obj.active
cell_obj = sheet_obj.cell(row =4 , column = 1)
print(cell_obj.value)
# get all the rows
print(sheet_obj.max_row)
# get all the columns
print(sheet_obj.max_column)
# print all column names
for i in range(1, sheet_obj.max_column + 1):
    cell_obj = sheet_obj.cell(row =1 , column = i)
    print(cell_obj.value)
# print the first column value
for i in range(1, sheet_obj.max_row+ 1):
    cell_obj = sheet_obj.cell(row =i , column = 1)
    print(cell_obj.value)
# get a particular row value
max_col = sheet_obj.max_column
for i in range(1 , max_col + 1):
    cell_obj= sheet_obj.cell(row = 2 , column = i)
    print(cell_obj.value, end = " ")
